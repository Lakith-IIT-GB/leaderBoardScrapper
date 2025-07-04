import os
import pandas as pd
import warnings
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests
from pathlib import Path

def load_env_file(env_path=".env"):
    """Load environment variables from .env file"""
    env_vars = {}
    try:
        if os.path.exists(env_path):
            with open(env_path, 'r') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, value = line.split('=', 1)
                        env_vars[key.strip()] = value.strip()
        return env_vars
    except Exception as e:
        print(f"Warning: Could not load .env file: {e}")
        return {}

# Load environment variables
env_vars = load_env_file()

class HackerRankLeaderboardCLI:
    def __init__(self):
        # Load configuration from .env file
        self.search_keyword = env_vars.get('SEARCH_KEYWORD', 'hackerrank')
        self.user_agent = env_vars.get('USER_AGENT', 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.80 Safari/537.36')
        self.request_timeout = int(env_vars.get('REQUEST_TIMEOUT', '10'))
        self.offset_limit = int(env_vars.get('OFFSET_LIMIT', '100'))
        self.max_offset = int(env_vars.get('MAX_OFFSET', '1000'))
        
        # Ensure Leaderboards directory exists
        Path("Leaderboards").mkdir(exist_ok=True)
        
        # Suppress warnings
        warnings.filterwarnings('ignore')

    def display_menu(self):
        print("\n" + "="*60)
        print("          HACKERRANK LEADERBOARD SCRAPER")
        print("="*60)
        print("1. Generate Excel Sheets from Contest IDs")
        print("2. Combine Existing Excel Sheets")
        print("0. Exit")
        print("="*60)

    def get_user_choice(self):
        while True:
            try:
                choice = input("\nEnter your choice (0/1/2): ").strip()
                if choice in ['0', '1', '2']:
                    return int(choice)
                else:
                    print("Invalid choice! Please enter 0, 1, or 2.")
            except KeyboardInterrupt:
                print("\n\nExiting...")
                return 0
            except Exception:
                print("Invalid input! Please enter 0, 1, or 2.")

    def get_contest_ids(self):
        print("\nEnter HackerRank Contest IDs:")
        print("(Separate multiple IDs with commas)")
        print("Example: contest1, contest2, contest3")
        
        while True:
            inp = input("\nContest IDs: ").strip()
            if not inp:
                print("Please enter at least one contest ID!")
                continue
            
            try:
                contest_ids = [id.strip() for id in inp.split(',') if id.strip()]
                if not contest_ids:
                    print("No valid contest IDs entered! Please try again.")
                    continue
                return contest_ids
            except Exception as e:
                print(f"Error parsing input: {e}. Please try again.")

    def generateExcelSheet(self, name, df):
        """Generate Excel sheet with formatting"""
        # Sort the DataFrame
        if name == 'TotalHackerrankLeaderBoard' or name == 'CombinedLeaderboard':
            df = df.sort_values(by='Total Score', ascending=False)
        else:
            df = df.sort_values(by='Score', ascending=False)

        # Add rank after sorting
        df.insert(0, 'Rank', range(1, len(df) + 1))

        # Create Excel file
        filepath = Path(f'Leaderboards/{name}.xlsx')
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            self.apply_excel_formatting(writer.sheets['Sheet1'], df)
        
        print(f"✓ Generated: {filepath}")

    def apply_excel_formatting(self, worksheet, df):
        """Apply formatting to Excel worksheet"""
        # Define styles
        styles = {
            'header': {
                'font': Font(name='Arial', size=18, bold=True),
                'fill': PatternFill(start_color='00ADEAEA', end_color='00ADEAEA', fill_type='solid'),
            },
            'body': {
                'font': Font(name='Arial', size=14, bold=True),
                'fill': PatternFill(start_color='00C7ECEC', end_color='00C7ECEC', fill_type='solid'),
            },
            'common': {
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(bottom=Side(style='medium'))
            }
        }

        # Set column widths
        worksheet.column_dimensions['A'].width = 12  # Rank column
        for col in worksheet.columns:
            column = col[0].column_letter
            if column != 'A':
                worksheet.column_dimensions[column].width = 35

        # Set row height
        for row in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 25

        # Apply formatting
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.value = value
            self.apply_cell_style(cell, styles['header'], styles['common'])

        for row_num, row in enumerate(df.values, start=2):
            for col_num, value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                self.apply_cell_style(cell, styles['body'], styles['common'])

    @staticmethod
    def apply_cell_style(cell, specific_style, common_style):
        """Apply styles to a cell"""
        for style_dict in (specific_style, common_style):
            for attr, value in style_dict.items():
                setattr(cell, attr, value)

    def fetch_hackerrank_data(self, tracker_name):
        """Fetch data from HackerRank API"""
        data = []
        headers = {"User-agent": self.user_agent}

        print(f"  Fetching data for: {tracker_name}")
        
        for offset in range(0, self.max_offset, self.offset_limit):
            url = f'https://www.hackerrank.com/rest/contests/{tracker_name}/leaderboard?offset={offset}&limit={self.offset_limit}'
            try:
                response = requests.get(url, headers=headers, timeout=self.request_timeout)
                response.raise_for_status()
                json_data = response.json()

                if not json_data.get('models'):
                    break

                for item in json_data['models']:
                    data.append({
                        'Name': item['hacker'],
                        'Score': item['score']
                    })

                print(f"    Fetched {len(data)} entries so far...")

            except requests.RequestException as e:
                print(f"  ❌ Error fetching data for {tracker_name}: {str(e)}")
                return None

        print(f"  ✓ Total entries fetched: {len(data)}")
        return pd.DataFrame(data) if data else None

    def generate_sheets(self, tracker_names):
        """Generate Excel sheets for given contest IDs"""
        print(f"\nGenerating sheets for {len(tracker_names)} contest(s)...")
        
        all_participants = {}
        
        for idx, tracker_name in enumerate(tracker_names, 1):
            print(f"\n[{idx}/{len(tracker_names)}] Processing: {tracker_name}")
            
            df = self.fetch_hackerrank_data(tracker_name)
            if df is None:
                continue

            if df.empty:
                print(f"  ⚠️ Warning: {tracker_name} returned no data")
                continue

            # Update all_participants dictionary
            for _, row in df.iterrows():
                if row['Name'] not in all_participants:
                    all_participants[row['Name']] = {contest: 0 for contest in tracker_names}
                all_participants[row['Name']][tracker_name] = row['Score']

            self.generateExcelSheet(tracker_name, df)

        # Generate total leaderboard
        if all_participants:
            print("\nGenerating combined leaderboard...")
            self.generate_total_leaderboard(all_participants, tracker_names)
            print("\n✅ All sheets generated successfully!")
            print(f"📁 Files saved in: {Path('Leaderboards').absolute()}")
        else:
            print("\n❌ No data was fetched. Please check your contest IDs.")

    def generate_total_leaderboard(self, all_participants, tracker_names):
        """Generate total leaderboard combining all contests"""
        total_data = []
        for participant, scores in all_participants.items():
            row = {'Name': participant}
            row.update(scores)
            row['Total Score'] = sum(scores.values())
            total_data.append(row)

        df_total = pd.DataFrame(total_data)
        columns = ['Name'] + tracker_names + ['Total Score']
        df_total = df_total[columns]
        self.generateExcelSheet('TotalHackerrankLeaderBoard', df_total)

    def get_file_path(self, title, file_type="Excel"):
        """Get file path from user input"""
        print(f"\n{title}")
        print("Enter the full path to the file:")
        
        while True:
            file_path = input("File path: ").strip().strip('"').strip("'")
            if not file_path:
                print("Please enter a valid file path!")
                continue
            
            path_obj = Path(file_path)
            if not path_obj.exists():
                print(f"File not found: {file_path}")
                print("Please check the path and try again.")
                continue
            
            if file_type == "Excel" and not file_path.lower().endswith('.xlsx'):
                print("Please select an Excel (.xlsx) file!")
                continue
            
            return file_path

    def combine_excel_sheets(self):
        """Combine student data with HackerRank leaderboard"""
        print("\n" + "="*60)
        print("           COMBINE EXCEL SHEETS")
        print("="*60)
        print("\nInstructions:")
        print("1. Provide Student Excel sheet with:")
        print("   - Option A: 'Roll number' and 'Hackerrank' columns")
        print("   - Option B: Just 'Hackerrank' usernames (auto-generates User_001, etc.)")
        print("   - Option C: Usernames in first column (any column name)")
        print("2. Provide HackerRank Leaderboard Excel sheet")
        print("   (typically 'TotalHackerrankLeaderBoard.xlsx')")

        try:
            # Get student file
            student_file = self.get_file_path("📁 Select Student Data Excel File")
            
            # Get HackerRank file
            hackerrank_file = self.get_file_path("📁 Select HackerRank Leaderboard Excel File")
            
            print("\nProcessing files...")
            
            # Read student data file
            print("📖 Reading student data file...")
            student_df = pd.read_excel(student_file)
            
            # Check if Roll number column exists, if not create a placeholder
            if 'Roll number' not in student_df.columns and 'Hackerrank' in student_df.columns:
                # Only usernames provided, create sequential roll numbers
                student_df['Roll number'] = [f"User_{i+1:03d}" for i in range(len(student_df))]
                student_df = student_df[['Roll number', 'Hackerrank']].copy()
            elif 'Roll number' in student_df.columns and 'Hackerrank' in student_df.columns:
                # Both columns exist
                student_df = student_df[['Roll number', 'Hackerrank']].copy()
            else:
                # Try to detect column names
                cols = student_df.columns.tolist()
                if len(cols) >= 1:
                    # Assume first column is usernames
                    student_df.columns = ['Hackerrank'] + [f'Col_{i}' for i in range(1, len(cols))]
                    student_df['Roll number'] = [f"User_{i+1:03d}" for i in range(len(student_df))]
                    student_df = student_df[['Roll number', 'Hackerrank']].copy()
                else:
                    raise ValueError("Could not identify username column in student file")

            student_df['Hackerrank'] = student_df['Hackerrank'].str.strip().str.lstrip('@').str.lower()

            # Read Hackerrank leaderboard file
            print("📖 Reading HackerRank leaderboard file...")
            hackerrank_df = pd.read_excel(hackerrank_file)

            # If there's a Rank column, drop it (we'll recreate it)
            if 'Rank' in hackerrank_df.columns:
                hackerrank_df = hackerrank_df.drop('Rank', axis=1)

            # Clean data and convert to lowercase for matching
            print("🔄 Processing and matching data...")
            student_df['Hackerrank'] = student_df['Hackerrank'].str.strip()
            hackerrank_df['Name_lower'] = hackerrank_df['Name'].str.strip().str.lower()

            # Merge dataframes using lowercase versions for matching
            result_df = pd.merge(
                hackerrank_df,
                student_df,
                left_on='Name_lower',
                right_on='Hackerrank',
                how='left'
            )

            # Drop the temporary and redundant columns
            result_df = result_df.drop(['Hackerrank', 'Name_lower'], axis=1)

            # Sort by Total Score if it exists, otherwise by Score
            sort_column = 'Total Score' if 'Total Score' in result_df.columns else 'Score'
            result_df = result_df.sort_values(sort_column, ascending=False)

            # Add Rank column and reorder columns
            result_df.insert(0, 'Rank', range(1, len(result_df) + 1))

            # Reorder columns to have Rank, Roll number, Name at the start
            cols = result_df.columns.tolist()
            cols.remove('Rank')
            cols.remove('Roll number')
            cols.remove('Name')
            final_cols = ['Rank', 'Roll number', 'Name'] + cols
            result_df = result_df[final_cols]

            # Generate Excel file
            print("📝 Generating combined Excel file...")
            output_path = Path('Leaderboards/CombinedLeaderboard.xlsx')
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Sheet1')
                self.apply_excel_formatting(writer.sheets['Sheet1'], result_df)

            print("\n✅ Combined Excel sheet generated successfully!")
            print(f"📁 File saved: {output_path.absolute()}")
            print(f"📊 Total entries: {len(result_df)}")
            print(f"👤 Matched entries (with Roll numbers): {result_df['Roll number'].notna().sum()}")
            print(f"❓ Unmatched entries: {result_df['Roll number'].isna().sum()}")

        except Exception as e:
            print(f"\n❌ Error occurred: {str(e)}")

    def run(self):
        """Main application loop"""
        print(f"🔧 Configuration loaded from .env:")
        print(f"  - Search keyword: {self.search_keyword}")
        print(f"  - Request timeout: {self.request_timeout}s")
        print(f"  - Offset limit: {self.offset_limit}")
        
        while True:
            self.display_menu()
            choice = self.get_user_choice()
            
            if choice == 0:
                print("\n👋 Goodbye!")
                break
            elif choice == 1:
                contest_ids = self.get_contest_ids()
                self.generate_sheets(contest_ids)
            elif choice == 2:
                self.combine_excel_sheets()
            
            input("\nPress Enter to continue...")


if __name__ == "__main__":
    try:
        app = HackerRankLeaderboardCLI()
        app.run()
    except KeyboardInterrupt:
        print("\n\n👋 Application terminated by user.")
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
