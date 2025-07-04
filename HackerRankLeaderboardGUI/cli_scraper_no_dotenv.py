import os
import pandas as pd
import warnings
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests
from pathlib import Path

class HackerRankLeaderboardCLI:
    def __init__(self):
        # Load configuration from .env file manually
        self.load_env_config()
        
        # Ensure Leaderboards directory exists
        Path("Leaderboards").mkdir(exist_ok=True)
        
        # Suppress warnings
        warnings.filterwarnings('ignore')

    def load_env_config(self):
        """Load configuration from .env file manually"""
        # Default values
        config = {
            'SEARCH_KEYWORD': 'hackerrank',
            'USER_AGENT': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.80 Safari/537.36',
            'REQUEST_TIMEOUT': '10',
            'OFFSET_LIMIT': '100',
            'MAX_OFFSET': '1000'
        }
        
        # Try to read .env file
        env_file = Path('.env')
        if env_file.exists():
            try:
                with open(env_file, 'r') as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith('#') and '=' in line:
                            key, value = line.split('=', 1)
                            config[key.strip()] = value.strip()
            except Exception as e:
                print(f"Warning: Could not read .env file: {e}")
        
        # Set configuration
        self.search_keyword = config['SEARCH_KEYWORD']
        self.user_agent = config['USER_AGENT']
        self.request_timeout = int(config['REQUEST_TIMEOUT'])
        self.offset_limit = int(config['OFFSET_LIMIT'])
        self.max_offset = int(config['MAX_OFFSET'])



    def generateExcelSheet(self, name, df):
        """Generate Excel sheet with formatting"""
        # Sort the DataFrame
        if name == 'TotalHackerrankLeaderBoard' or name == 'CombinedLeaderboard':
            df = df.sort_values(by='Total Score', ascending=False)
        else:
            df = df.sort_values(by='Score', ascending=False)

        # Add rank after sorting
        df.insert(0, 'Rank', range(1, len(df) + 1))

        # Create Excel file
        filepath = Path(f'Leaderboards/{name}.xlsx')
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            self.apply_excel_formatting(writer.sheets['Sheet1'], df)
        
        print(f"✓ Generated: {filepath}")

    def apply_excel_formatting(self, worksheet, df):
        """Apply formatting to Excel worksheet"""
        # Define styles
        styles = {
            'header': {
                'font': Font(name='Arial', size=18, bold=True),
                'fill': PatternFill(start_color='00ADEAEA', end_color='00ADEAEA', fill_type='solid'),
            },
            'body': {
                'font': Font(name='Arial', size=14, bold=True),
                'fill': PatternFill(start_color='00C7ECEC', end_color='00C7ECEC', fill_type='solid'),
            },
            'common': {
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(bottom=Side(style='medium'))
            }
        }

        # Set column widths
        worksheet.column_dimensions['A'].width = 12  # Rank column
        for col in worksheet.columns:
            column = col[0].column_letter
            if column != 'A':
                worksheet.column_dimensions[column].width = 35

        # Set row height
        for row in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 25

        # Apply formatting
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.value = value
            self.apply_cell_style(cell, styles['header'], styles['common'])

        for row_num, row in enumerate(df.values, start=2):
            for col_num, value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                self.apply_cell_style(cell, styles['body'], styles['common'])

    @staticmethod
    def apply_cell_style(cell, specific_style, common_style):
        """Apply styles to a cell"""
        for style_dict in (specific_style, common_style):
            for attr, value in style_dict.items():
                setattr(cell, attr, value)

    def fetch_hackerrank_data(self, tracker_name):
        """Fetch data from HackerRank API"""
        data = []
        headers = {"User-agent": self.user_agent}

        print(f"  Fetching data for: {tracker_name}")
        
        for offset in range(0, self.max_offset, self.offset_limit):
            url = f'https://www.hackerrank.com/rest/contests/{tracker_name}/leaderboard?offset={offset}&limit={self.offset_limit}'
            try:
                response = requests.get(url, headers=headers, timeout=self.request_timeout)
                response.raise_for_status()
                json_data = response.json()

                if not json_data.get('models'):
                    break

                for item in json_data['models']:
                    data.append({
                        'Name': item['hacker'],
                        'Score': item['score']
                    })

                print(f"    Fetched {len(data)} entries so far...")

            except requests.RequestException as e:
                print(f"  ❌ Error fetching data for {tracker_name}: {str(e)}")
                return None

        print(f"  ✓ Total entries fetched: {len(data)}")
        return pd.DataFrame(data) if data else None

    def generate_sheets(self, tracker_names):
        """Generate Excel sheets for given contest IDs"""
        print(f"\nGenerating sheets for {len(tracker_names)} contest(s)...")
        
        all_participants = {}
        
        for idx, tracker_name in enumerate(tracker_names, 1):
            print(f"\n[{idx}/{len(tracker_names)}] Processing: {tracker_name}")
            
            df = self.fetch_hackerrank_data(tracker_name)
            if df is None:
                continue

            if df.empty:
                print(f"  ⚠️ Warning: {tracker_name} returned no data")
                continue

            # Update all_participants dictionary
            for _, row in df.iterrows():
                if row['Name'] not in all_participants:
                    all_participants[row['Name']] = {contest: 0 for contest in tracker_names}
                all_participants[row['Name']][tracker_name] = row['Score']

            self.generateExcelSheet(tracker_name, df)

        # Generate total leaderboard
        if all_participants:
            print("\nGenerating combined leaderboard...")
            self.generate_total_leaderboard(all_participants, tracker_names)
            print("\n✅ All sheets generated successfully!")
            print(f"📁 Files saved in: {Path('Leaderboards').absolute()}")
        else:
            print("\n❌ No data was fetched. Please check your contest IDs.")

    def generate_total_leaderboard(self, all_participants, tracker_names):
        """Generate total leaderboard combining all contests"""
        total_data = []
        for participant, scores in all_participants.items():
            row = {'Name': participant}
            row.update(scores)
            row['Total Score'] = sum(scores.values())
            total_data.append(row)

        df_total = pd.DataFrame(total_data)
        columns = ['Name'] + tracker_names + ['Total Score']
        df_total = df_total[columns]
        self.generateExcelSheet('TotalHackerrankLeaderBoard', df_total)

    def run(self):
        """Main application loop"""
        print("\n" + "="*60)
        print("          HACKERRANK LEADERBOARD SCRAPER")
        print("="*60)
        print(f"🔧 Configuration loaded:")
        print(f"  - Search keyword: {self.search_keyword}")
        print(f"  - Request timeout: {self.request_timeout}s")
        print(f"  - Offset limit: {self.offset_limit}")
        
        print(f"\n📊 Processing contest: coderally-6-0-training-weeks")
        contest_ids = ["coderally-6-0-training-weeks"]
        self.generate_sheets(contest_ids)
        
        print(f"\n✅ Process completed! Check the Leaderboards folder for results.")


if __name__ == "__main__":
    try:
        app = HackerRankLeaderboardCLI()
        app.run()
    except KeyboardInterrupt:
        print("\n\n👋 Application terminated by user.")
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
